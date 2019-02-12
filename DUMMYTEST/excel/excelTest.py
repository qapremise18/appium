import openpyxl
import sys
from openpyxl import load_workbook
from openpyxl import utils


class ExcelTest:
# wb = openpyxl.load_workbook('D:\\PythonWS\\Premise\\resources\\PremiseTestData.xlsx')
# print(wb.sheetnames)
# print(type(wb))
# wb.sheetnames.__getitem__("")


    def getDataInHashMap(self , dataSheetFilePath , sheetName, dataKey):
        testData = {}
        try:
            # testDataArray = {}
            testDataArray = self.readExcelData(dataSheetFilePath, sheetName, dataKey)
            for i in range(len(testDataArray)) :
                testData[testDataArray[i][0]]= testDataArray[i][1]
                # testData[testDataArray[i][0]] =  testDataArray[i][1]
        except :
            print(" Failed to read excel data by data key  and store in linked hash map due to :::" + sys.exc_info())
        return testData


    def readExcelData(self, filePath, sheetName, tableName):
        testData = None
        print("Reading file, sheet, table::: " , filePath + ", " , sheetName + ", " , tableName)
        try:
            workbook = openpyxl.load_workbook(filePath)
            sheet = workbook.get_sheet_by_name(sheetName)
            boundaryCells = self.findCell(sheet, tableName)
            startCell = boundaryCells[0]
            endCell = boundaryCells[1]
            print("readExcelData start end cell>>",startCell,"==",endCell)
            print("type>>@@@",type(startCell),"@@",type(endCell))
            startRow = startCell + 1
            endRow = boundaryCells[2]

            startCol = endCell + 1
            sheet_obj = workbook.active

            endCol = boundaryCells[3] - 1
            # testData = [] # String[endRow - startRow + 1][endCol - startCol + 1]
            print("@@@TESTstartRow",startRow,"==endRow",endRow)
            print("@@@TESTstartCol", startCol, "==endCol", endCol)
            testData = [[0] * endCol for i in range(endRow-startRow+1)]
            print("len>>",len(testData))
            for i in range(startRow, endRow+1):
                print("for>>")
                for j in range(startCol,endCol+1):
                    testData[i - startRow][j - startCol]= sheet_obj.cell(row=i,column=j).value
                    print("for>>22")
            print("PRRRRR>>>",testData)
        except:
            print("EORRRR readExcelData" , sys.exc_info())
        return testData


    def findCell(self, sheet, text):
        pos = "start"
        cells = []
        # get max row count
        max_row = sheet.max_row +1
        # get max column count
        max_column = sheet.max_column +1
        min_row =1
        #for row in sheet.iter_rows(min_row , max_column, max_row):
        print("PPPP>>",max_row,"==",max_column)

        try:
            for row in sheet.iter_rows(min_row=1, max_col=max_column, max_row=max_row):
                for cell in row:
                    print("for each cell>>>",cell.value)
                    if text == cell.value :
                        print("In Cell val eyal to textXX>>",cell.row,"==",cell.column)
                        if pos == "start":
                            print("In Cell val start")
                            # cells[0] = cell
                            cells.append(cell.row)
                            cells.append(cell.column)
                            pos = "end"
                        else:
                            print("In Cell val end")
                            # cells[1] = cell
                            cells.append(cell.row)
                            cells.append(cell.column)
            print("HHHHH>>",cells,"==",cells[0],"==",type(cells[0]))
            # print(sheet["A47"])
            # print(openpyxl.utils.cell.get_column_letter("A47"))
            # print(cell.row,"=====RCC=",cell.column)
            return cells
        except:
            print("Erorr in find cell>>>",sys.exc_info())

    def roww(self):
        wb = openpyxl.load_workbook('D:\\PythonWS\\Premise\\resources\\PremiseTestData.xlsx')
        print(wb.sheetnames)
        sheet = wb.active
        for row in wb.iter_rows(min_row=1, max_col=3, max_row=2):
            for cell in row:
                print(cell)

    def excel_test(self):
        # filepath = "D:\\PythonWS\\Premise\\resources\\PremiseTestData.xlsx"

        filepath = "K:\\PythonWS\\Premise\\resources\\PremiseTestData.xlsx"
        # load demo.xlsx

        wb = openpyxl.load_workbook(filepath)
        # select demo.xlsx
        #
        sheet = wb.get_sheet_by_name("Smoke")
        # get max row count
        max_row = sheet.max_row
        # get max column count
        max_column = sheet.max_column
        # iterate over all cells
        # iterate over all rows
        for i in range(3, max_row + 1):

            # iterate over all columns
            for j in range(1, max_column + 1):
                # get particular cell value
                cell_obj = sheet.cell(row=i, column=j)
                # print("####",type(cell_obj))
                # print cell value
                print(cell_obj.value, end=' | ')
                # print("####", type(cell_obj.value))
            # print new line
            print('\n')
        self.findCell(sheet,"PremiseTests")

# filePath = 'D:\\PythonWS\\Premise\\resources\\PremiseTestData.xlsx'
filepath = "D:\\PythonWS\\Premise\\resources\\PremiseTestData.xlsx"
test = ExcelTest()
# test.getDataInHashMap(filePath,)
# test.excel_test()

# arr = test.readExcelData(filepath,"Smoke","PremiseTests")
# print("########################")
# filepath = "D:\\PythonWS\\Premise\\resources\\PremiseTestData.xlsx"
# test = ExcelTest()
# arr = test.readExcelData(filepath,"login","LoginTestData")
# map = {}
# print("ARR",arr[0][1])
# print("ZZZ",type(arr))
# print("CCCC",len(arr))
# for i in range(len(arr)):
#     print("XXXXXXXxx")
#     map[arr[i][0]]= arr[i][1]
# print("map",map["HaveCodeTestDataDreams"])
# print("########################")

dic = test.getDataInHashMap(filepath,"login", "Locators")
print("$$",type(dic))
print("$$",len(dic))
print("##",dic["JoinNowButton"])
# test.excel_test()